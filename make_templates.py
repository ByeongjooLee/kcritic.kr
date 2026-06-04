"""
make_templates.py
두 개의 Excel 템플릿 파일을 생성합니다.
  - templates/template_essay.xlsx    (비평글 입력 양식)
  - templates/template_research.xlsx (선행연구 입력 양식)
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

# ── 공통 스타일 ──────────────────────────────────────────────
HEADER_BG = "DDEEFF"
HEADER_FILL = PatternFill(start_color=HEADER_BG, end_color=HEADER_BG, fill_type="solid")
HEADER_FONT = Font(bold=True)

def style_header(ws, col_idx):
    cell = ws.cell(row=1, column=col_idx)
    cell.fill = HEADER_FILL
    cell.font = HEADER_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center")


# ── 파일 1: template_essay.xlsx ──────────────────────────────
def make_essay_template(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "비평글"

    # 헤더 정의: (컬럼명, 필수여부, 너비)
    columns = [
        ("비평가",        True,  12),
        ("에세이_제목",   True,  30),
        ("발표지",        True,  20),
        ("연도",          True,   8),
        ("비평_대상_작가", True,  15),
        ("인용_이론가",   False, 25),
        ("출처_URL",      False, 30),
    ]

    for col_idx, (name, required, width) in enumerate(columns, start=1):
        label = f"{name}*" if required else name
        ws.cell(row=1, column=col_idx, value=label)
        style_header(ws, col_idx)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 예시 데이터 (2행)
    sample = [
        "김우창",
        "시적 인간과 사회 현실",
        "세계의 문학",
        1979,
        "김지하",
        "마르쿠제, 아도르노",
        "https://",
    ]
    for col_idx, value in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    # 행 높이
    ws.row_dimensions[1].height = 20

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"[OK] {path}")


# ── 파일 2: template_research.xlsx ──────────────────────────
def make_research_template(path: str):
    wb = Workbook()
    ws = wb.active
    ws.title = "선행연구"

    # 헤더 정의: (컬럼명, 필수여부, 너비)
    columns = [
        ("논문_제목",   True,  35),
        ("저자",        True,  12),
        ("대학_기관",   True,  20),
        ("연도",        True,   8),
        ("종류",        True,  10),
        ("대상_작가",   False, 15),
        ("장르",        False, 10),
        ("지도교수",    False, 12),
        ("주제어",      False, 30),
    ]

    for col_idx, (name, required, width) in enumerate(columns, start=1):
        label = f"{name}*" if required else name
        ws.cell(row=1, column=col_idx, value=label)
        style_header(ws, col_idx)
        ws.column_dimensions[get_column_letter(col_idx)].width = width

    # 예시 데이터 (2행)
    sample = [
        "윤동주 시의 자아 연구",
        "김용주",
        "국민대학교",
        2004,
        "phd",
        "윤동주",
        "시",
        "홍길동",
        "자아, 저항, 식민지",
    ]
    for col_idx, value in enumerate(sample, start=1):
        ws.cell(row=2, column=col_idx, value=value)

    # 종류 컬럼(E열)에 드롭다운 유효성 검사 (3행~1000행)
    dv = DataValidation(
        type="list",
        formula1='"phd,master,kci"',
        allow_blank=True,
        showDropDown=False,   # False = 드롭다운 화살표 표시
    )
    dv.sqref = "E3:E1000"
    dv.error = "phd, master, kci 중 하나를 선택하세요."
    dv.errorTitle = "입력 오류"
    dv.prompt = "phd / master / kci"
    dv.promptTitle = "종류 선택"
    ws.add_data_validation(dv)

    # 행 높이
    ws.row_dimensions[1].height = 20

    os.makedirs(os.path.dirname(path), exist_ok=True)
    wb.save(path)
    print(f"[OK] {path}")


# ── 실행 ─────────────────────────────────────────────────────
if __name__ == "__main__":
    base = os.path.dirname(os.path.abspath(__file__))
    make_essay_template(os.path.join(base, "templates", "template_essay.xlsx"))
    make_research_template(os.path.join(base, "templates", "template_research.xlsx"))
    print("완료: templates/ 디렉터리에 두 파일이 생성되었습니다.")
